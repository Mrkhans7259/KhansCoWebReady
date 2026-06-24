import sqlite3
from openpyxl import load_workbook

DB_PATH = "khansco.db"


class ImportRepository:

    def __init__(self):
        self.conn = sqlite3.connect(DB_PATH)
        self.cursor = self.conn.cursor()

    def import_clients_from_excel(self, file_path):
        workbook = load_workbook(file_path)
        sheet = workbook.active

        imported = 0
        updated = 0

        for row in sheet.iter_rows(min_row=2, values_only=True):
            client_code, client_name, gstin, mobile, email, assigned_staff = row

            if not client_code or not client_name:
                continue

            self.cursor.execute(
                "SELECT id FROM clients WHERE client_code = ?",
                (str(client_code),)
            )
            existing = self.cursor.fetchone()

            if existing:
                self.cursor.execute("""
                    UPDATE clients
                    SET client_name = ?, gstin = ?, mobile = ?, email = ?, assigned_staff = ?
                    WHERE client_code = ?
                """, (
                    client_name,
                    gstin or "",
                    mobile or "",
                    email or "",
                    assigned_staff or "",
                    str(client_code)
                ))
                updated += 1
            else:
                self.cursor.execute("""
                    INSERT INTO clients
                    (client_code, client_name, gstin, mobile, email, assigned_staff)
                    VALUES (?, ?, ?, ?, ?, ?)
                """, (
                    str(client_code),
                    client_name,
                    gstin or "",
                    mobile or "",
                    email or "",
                    assigned_staff or ""
                ))
                imported += 1

        self.conn.commit()

        return imported, updated